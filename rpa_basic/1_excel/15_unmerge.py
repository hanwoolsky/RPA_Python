from openpyxl import load_workbook
wb = load_workbook("sample_merge.xlsx")
ws = wb.active

# 셀 나누기
ws.unmerge_cells("B2:D2") #B2부터 D2까지 합치겠음
wb.save("sample_merge.xlsx")