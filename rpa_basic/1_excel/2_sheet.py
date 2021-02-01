from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() # 새로운 Sheet 기본 이름으로 생성 (시트 추가)
ws.title = "Mysheet" # Sheet 이름 변경
ws.sheet_properties.tabColor =  "ff66ff" # RGB 형태로 값을 넣어주면 아래 쪽 탭 색상 변경    w3schools.com/colors

ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet("NewSheet", 2) # 2번째 index에 Sheet 생성

new_ws = wb["NewSheet"] # Dict 형태로 Sheet에 접근 가능

print(wb.sheetnames) # 모든 Sheet 이름이 출력

# Sheet 복사
new_ws["A1"] = "Test" # A1칸에 데이터 추가
target = wb.copy_worksheet(new_ws) # new_ws Sheet 복사
target.title = "Copied Sheet"

wb.save("sample.xlsx")