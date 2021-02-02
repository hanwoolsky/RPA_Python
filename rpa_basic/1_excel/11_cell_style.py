from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
wb = load_workbook("sample.xlsx")
ws = wb.active

# 번호, 영어, 수학
a1 = ws["A1"]
b1 = ws["B1"]
c1 = ws["C1"]

ws.column_dimensions["A"].width = 5 # A열의 너비를 5로 설정
ws.row_dimensions[1].height = 50 # 1열의 높이를 50으로 설정

# 스타일 적용
a1.font = Font(color = "FF0000", italic = True, bold = True) # 글자 색은 빨강, 기울임 + 두껍게 적용
b1.font = Font(color = "CC33FF", name = "Arial", strike = True) # 폰트 지정 + 취소선
c1.font = Font(color = "0000FF", size = 20, underline = "single") # 글자 크기 지정 + 밑줄

thin_border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin")) # 테두리 적용
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 90점 넘는 셀은 초록색
for row in ws.rows:
    for cell in row:
        cell.alignment = Alignment(horizontal = "center", vertical = "center") # 각 cell에 대해 중앙 정렬 (center, left, right, top, bottom)
        if cell.column == 1: # A: 번호 열은 제외
            continue
        if isinstance(cell.value, int) and cell.value > 90: # 정수형 데이터이면 (영어, 수학 문자는 제외)
            cell.fill = PatternFill(fgColor = "00FF00", fill_type = "solid") # 배경색을 초록으로
            cell.font = Font(color = "FF0000") # 글자는 빨강

# title 고정 (스크롤해도 계속 보이게)
ws.freeze_panes = "B2" # B2 기준으로 왼쪽, 위 고정 (A열이 계속 보이고 1번째 row가 계속 보인다.)

wb.save("sample_style.xlsx")