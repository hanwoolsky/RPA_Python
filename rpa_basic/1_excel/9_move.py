from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

# 번호 영어 수학
# 번호 (국어) 영어 수학

ws.move_range("B1:C11", rows = 0, cols = 1) #B1에서 C11의 데이터를 열로만 오른쪽으로 이동(-1은 왼쪽으로)
ws["B1"].value = "국어" # B 셀에 '국어' 입력

wb.save("sample_korean.xlsx")