from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

# 1줄 씩 데이터 넣기
ws.append(["번호", "영어", "수학"]) # A, B, C
for i in range(1, 11): # 10개 데이터 넣기
    ws.append([i, randint(0,100), randint(0,100)])

col_B = ws["B"] # 영어 column만 가져오기

for cell in col_B:
    print(cell.value)

col_range = ws["B:C"] # 영어, 수학 column 함께 가져오기
for cols in col_range:
    for cell in cols:
        print(cell.value)

row_title = ws[1]
for cell in row_title:
    print(cell.value)

row_range = ws[2:6] # 2번째 줄에서 6번째 줄까지 가지고 오기 (1번째 title 제외)
for rows in row_range:
    for cell in rows:
        print(cell.value, end = " ")
    print()

from openpyxl.utils.cell import coordinate_from_string

row_range = ws[2:ws.max_row] # 2번째 줄부터 마지막 줄까지
for rows in row_range:
    for cell in rows:
        # print(cell.value, end = " ")
        print(cell.coordinate, end = " ") # A2, B4와 같은 셀의 좌표 정보를 출력
        xy = coordinate_from_string(cell.coordinate) # ('A', 3) 으로 끊어줌
        print(xy, end = " ")
        print(xy[0], end = " ")
        print(xy[1], end = " ") 
    print()

# 전체 rows
print(tuple(ws.rows))
for row in tuple(ws.rows):
    print(row[1].value)

# 전체 columns
print(tuple(ws.columns))
for column in tuple(ws.columns):
    print(column[0].value)

for row in ws.iter_rows(): # 전체 row
    print(row[2].value)

for column in ws.iter_columns(): # 전체 column
    print(column[2].value)

# 1번째 줄부터 5번째 줄까지 2번째 열부터 3번째 열까지
for row in ws.iter_rows(min_row = 1, max_row = 5, min_col = 2, max_col = 3):
    print(row[2].value)
    print(row) # 데이터를 좌우좌우로 가져옴

for col in ws.iter_cols(min_row = 1, max_row = 5, min_col = 2, max_col = 3):
    print(col) # 데이터를 상하상하로 가져옴

wb.save("sample.xlsx")