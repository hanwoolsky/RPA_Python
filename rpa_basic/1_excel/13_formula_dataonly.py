from openpyxl import load_workbook
wb = load_workbook("sample_formula.xlsx")
ws = wb.active

# 수식이 그대로 뜸
for row in ws.values:
    for cell in row:
        print(cell)

wb = load_workbook("sample_formula.xlsx", data_only = True) # 수식이 아닌 실제 데이터를 가지고 옴
ws = wb.active

for row in ws.values:
    for cell in row:
        print(cell) # None으로 뜨는 것은 수식이 계산되지 않아서! 엑셀을 한번 열었다가 저장하고 닫으면 데이터를 받아올 수 있음